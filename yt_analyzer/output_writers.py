"""
output_writers.py — one writer class per --format, all sharing a common
streaming/chunked contract so multiple formats can be written
simultaneously (--format json,csv,txt,xlsx,db) without ever holding the
full result set in memory at once.

Contract (BaseWriter):
    open()               — acquire file handles / write headers
    write_batch(records)  — append one chunk (called once per flush cycle,
                             e.g. every N videos per the blueprint's
                             chunking requirement)
    close()               — flush + finalize (write closing brackets,
                             save the workbook, etc.)

Why each format needs a different chunking strategy:
    * JSONL:  trivially streamable — one JSON object per line, just append.
    * JSON:   a single valid JSON array requires care: we stream `[`, then
              comma-separated objects across batches, then `]` at close();
              we never build the whole array in memory.
    * CSV:    streamable via csv.DictWriter in append mode; header written
              once on first batch. Nested data (comments) is flattened via
              one row per comment (a "long" format) in a companion file,
              since CSV has no native nested-object representation.
    * XLSX:   openpyxl's write-only mode streams rows to disk incrementally
              instead of building the whole workbook in memory, which
              matters once "millions of comments" (per the blueprint) are
              in play.
    * TXT:    a human/LLM-readable structured block per video, appended
              incrementally — this is the "structured for LLM readability"
              format called for in the blueprint, distinct from raw JSON.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Optional, TextIO

from .exceptions import OutputWriterError
from .models import VideoRecord

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    _OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover - guarded import; requirements.txt pins this, but fail soft
    _OPENPYXL_AVAILABLE = False


# --------------------------------------------------------------------------- #
# Base contract
# --------------------------------------------------------------------------- #
class BaseWriter:
    format_name: str = "base"

    def __init__(self, output_path: str | Path, fields: Optional[list[str]] = None):
        self.output_path = Path(output_path)
        self.fields = fields
        self._opened = False

    def open(self) -> None:
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self._opened = True

    def write_batch(self, records: list[VideoRecord]) -> None:
        raise NotImplementedError

    def close(self) -> None:
        self._opened = False

    def __enter__(self) -> "BaseWriter":
        self.open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.close()


# --------------------------------------------------------------------------- #
# JSONL — JSON Lines, ideal for vector-DB / streaming ingestion pipelines
# --------------------------------------------------------------------------- #
class JSONLWriter(BaseWriter):
    format_name = "jsonl"

    def __init__(self, output_path: str | Path, fields: Optional[list[str]] = None):
        super().__init__(output_path, fields)
        self._fh: Optional[TextIO] = None

    def open(self) -> None:
        super().open()
        try:
            self._fh = self.output_path.open("w", encoding="utf-8")
        except OSError as exc:
            raise OutputWriterError(f"Could not open JSONL output '{self.output_path}': {exc}") from exc

    def write_batch(self, records: list[VideoRecord]) -> None:
        if self._fh is None:
            raise OutputWriterError("JSONLWriter.write_batch called before open()")
        for record in records:
            line = json.dumps(record.to_dict(self.fields), ensure_ascii=False)
            self._fh.write(line + "\n")
        self._fh.flush()

    def close(self) -> None:
        if self._fh is not None:
            self._fh.close()
            self._fh = None
        super().close()


# --------------------------------------------------------------------------- #
# JSON — a single valid JSON array, streamed incrementally
# --------------------------------------------------------------------------- #
class JSONWriter(BaseWriter):
    format_name = "json"

    def __init__(self, output_path: str | Path, fields: Optional[list[str]] = None):
        super().__init__(output_path, fields)
        self._fh: Optional[TextIO] = None
        self._wrote_any = False

    def open(self) -> None:
        super().open()
        try:
            self._fh = self.output_path.open("w", encoding="utf-8")
            self._fh.write("[\n")
        except OSError as exc:
            raise OutputWriterError(f"Could not open JSON output '{self.output_path}': {exc}") from exc

    def write_batch(self, records: list[VideoRecord]) -> None:
        if self._fh is None:
            raise OutputWriterError("JSONWriter.write_batch called before open()")
        for record in records:
            if self._wrote_any:
                self._fh.write(",\n")
            line = json.dumps(record.to_dict(self.fields), ensure_ascii=False, indent=2)
            self._fh.write(line)
            self._wrote_any = True
        self._fh.flush()

    def close(self) -> None:
        if self._fh is not None:
            self._fh.write("\n]\n")
            self._fh.close()
            self._fh = None
        super().close()


# --------------------------------------------------------------------------- #
# CSV — flat rows; comments (if selected) go to a companion *_comments.csv
# --------------------------------------------------------------------------- #
class CSVWriter(BaseWriter):
    format_name = "csv"

    _COMMENT_FIELDNAMES = (
        "video_id", "comment_id", "parent_id", "author", "author_channel_id",
        "text", "like_count", "reply_count", "is_favorited", "is_pinned", "timestamp",
    )
    _CHAPTER_FIELDNAMES = ("video_id", "chapter_index", "start_time", "end_time", "title")

    def __init__(self, output_path: str | Path, fields: Optional[list[str]] = None):
        super().__init__(output_path, fields)
        self._fh: Optional[TextIO] = None
        self._writer: Optional[csv.DictWriter] = None
        self._comments_fh: Optional[TextIO] = None
        self._comments_writer: Optional[csv.DictWriter] = None
        self._want_comments = fields is None or "comments" in fields
        self._comments_path = self.output_path.with_name(
            self.output_path.stem + "_comments" + self.output_path.suffix
        )
        self._chapters_fh: Optional[TextIO] = None
        self._chapters_writer: Optional[csv.DictWriter] = None
        self._want_chapters = fields is None or "chapters" in fields
        self._chapters_path = self.output_path.with_name(
            self.output_path.stem + "_chapters" + self.output_path.suffix
        )

    def open(self) -> None:
        super().open()
        try:
            self._fh = self.output_path.open("w", encoding="utf-8", newline="")
        except OSError as exc:
            raise OutputWriterError(f"Could not open CSV output '{self.output_path}': {exc}") from exc
        if self._want_comments:
            try:
                self._comments_fh = self._comments_path.open("w", encoding="utf-8", newline="")
                self._comments_writer = csv.DictWriter(self._comments_fh, fieldnames=self._COMMENT_FIELDNAMES)
                self._comments_writer.writeheader()
            except OSError as exc:
                raise OutputWriterError(f"Could not open companion CSV '{self._comments_path}': {exc}") from exc
        if self._want_chapters:
            try:
                self._chapters_fh = self._chapters_path.open("w", encoding="utf-8", newline="")
                self._chapters_writer = csv.DictWriter(self._chapters_fh, fieldnames=self._CHAPTER_FIELDNAMES)
                self._chapters_writer.writeheader()
            except OSError as exc:
                raise OutputWriterError(f"Could not open companion CSV '{self._chapters_path}': {exc}") from exc

    def _flat_dict_for_csv(self, record: VideoRecord) -> dict[str, Any]:
        d = record.to_dict(self.fields)
        # CSV cannot hold nested structures — drop the raw comment/chapter
        # objects (they go to their own companion files) and json-encode any
        # remaining list/dict values (tags, categories) so csv.writer doesn't
        # choke on non-string values.
        d.pop("comments", None)
        d.pop("chapters", None)
        for key, value in list(d.items()):
            if isinstance(value, (list, dict)):
                d[key] = json.dumps(value, ensure_ascii=False)
        return d

    def write_batch(self, records: list[VideoRecord]) -> None:
        if self._fh is None:
            raise OutputWriterError("CSVWriter.write_batch called before open()")

        flat_rows = [self._flat_dict_for_csv(r) for r in records]
        if flat_rows and self._writer is None:
            # Header is determined from the first batch's keys; every
            # VideoRecord.to_dict() call with the same `fields` filter
            # produces the same key set, so this is stable across batches.
            fieldnames = list(flat_rows[0].keys())
            self._writer = csv.DictWriter(self._fh, fieldnames=fieldnames, extrasaction="ignore")
            self._writer.writeheader()
        if self._writer is not None:
            self._writer.writerows(flat_rows)
        self._fh.flush()

        if self._want_comments and self._comments_writer is not None:
            comment_rows = []
            for record in records:
                for top in record.comments:
                    for c in top.flatten():
                        comment_rows.append(
                            {
                                "video_id": record.video_id,
                                "comment_id": c.comment_id,
                                "parent_id": c.parent_id,
                                "author": c.author,
                                "author_channel_id": c.author_channel_id,
                                "text": c.text,
                                "like_count": c.like_count,
                                "reply_count": c.reply_count,
                                "is_favorited": c.is_favorited,
                                "is_pinned": c.is_pinned,
                                "timestamp": c.timestamp,
                            }
                        )
            if comment_rows:
                self._comments_writer.writerows(comment_rows)
                self._comments_fh.flush()

        if self._want_chapters and self._chapters_writer is not None:
            chapter_rows = []
            for record in records:
                for idx, ch in enumerate(record.chapters, start=1):
                    chapter_rows.append(
                        {
                            "video_id": record.video_id,
                            "chapter_index": idx,
                            "start_time": ch.start_time,
                            "end_time": ch.end_time,
                            "title": ch.title,
                        }
                    )
            if chapter_rows:
                self._chapters_writer.writerows(chapter_rows)
                self._chapters_fh.flush()

    def close(self) -> None:
        if self._fh is not None:
            self._fh.close()
            self._fh = None
        if self._comments_fh is not None:
            self._comments_fh.close()
            self._comments_fh = None
        if self._chapters_fh is not None:
            self._chapters_fh.close()
            self._chapters_fh = None
        super().close()


# --------------------------------------------------------------------------- #
# XLSX — streamed via openpyxl write-only mode (never builds full workbook in RAM)
# --------------------------------------------------------------------------- #
class XLSXWriter(BaseWriter):
    format_name = "xlsx"

    _COMMENT_FIELDNAMES = (
        "video_id", "comment_id", "parent_id", "author", "author_channel_id",
        "text", "like_count", "reply_count", "is_favorited", "is_pinned", "timestamp",
    )
    _CHAPTER_FIELDNAMES = ("video_id", "chapter_index", "start_time", "end_time", "title")

    def __init__(self, output_path: str | Path, fields: Optional[list[str]] = None):
        if not _OPENPYXL_AVAILABLE:
            raise OutputWriterError(
                "openpyxl is required for --format xlsx but is not installed. "
                "Install it with: pip install openpyxl"
            )
        super().__init__(output_path, fields)
        self._wb: Optional["Workbook"] = None
        self._ws_videos: Optional["Worksheet"] = None
        self._ws_comments: Optional["Worksheet"] = None
        self._ws_chapters: Optional["Worksheet"] = None
        self._header_written = False
        self._fieldnames: list[str] = []
        self._want_comments = fields is None or "comments" in fields
        self._want_chapters = fields is None or "chapters" in fields

    def open(self) -> None:
        super().open()
        # write_only=True is the key memory-management feature here: rows
        # are streamed straight to a temp file on disk as they're appended,
        # rather than accumulating as Python objects in memory.
        self._wb = Workbook(write_only=True)
        self._ws_videos = self._wb.create_sheet("Videos")
        if self._want_comments:
            self._ws_comments = self._wb.create_sheet("Comments")
            self._ws_comments.append(list(self._COMMENT_FIELDNAMES))
        if self._want_chapters:
            self._ws_chapters = self._wb.create_sheet("Chapters")
            self._ws_chapters.append(list(self._CHAPTER_FIELDNAMES))

    def _flat_row_for_xlsx(self, record: VideoRecord) -> dict[str, Any]:
        d = record.to_dict(self.fields)
        d.pop("comments", None)
        d.pop("chapters", None)
        for key, value in list(d.items()):
            if isinstance(value, (list, dict)):
                d[key] = json.dumps(value, ensure_ascii=False)
        return d

    def write_batch(self, records: list[VideoRecord]) -> None:
        if self._wb is None or self._ws_videos is None:
            raise OutputWriterError("XLSXWriter.write_batch called before open()")

        flat_rows = [self._flat_row_for_xlsx(r) for r in records]
        if flat_rows and not self._header_written:
            self._fieldnames = list(flat_rows[0].keys())
            self._ws_videos.append(self._fieldnames)
            self._header_written = True

        for row in flat_rows:
            self._ws_videos.append([row.get(k, "") for k in self._fieldnames])

        if self._want_comments and self._ws_comments is not None:
            for record in records:
                for top in record.comments:
                    for c in top.flatten():
                        self._ws_comments.append(
                            [
                                record.video_id, c.comment_id, c.parent_id, c.author,
                                c.author_channel_id, c.text, c.like_count, c.reply_count,
                                c.is_favorited, c.is_pinned, c.timestamp,
                            ]
                        )

        if self._want_chapters and self._ws_chapters is not None:
            for record in records:
                for idx, ch in enumerate(record.chapters, start=1):
                    self._ws_chapters.append(
                        [record.video_id, idx, ch.start_time, ch.end_time, ch.title]
                    )

    def close(self) -> None:
        if self._wb is not None:
            try:
                self._wb.save(str(self.output_path))
            except OSError as exc:
                raise OutputWriterError(f"Could not save XLSX workbook '{self.output_path}': {exc}") from exc
            finally:
                self._wb = None
        super().close()


# --------------------------------------------------------------------------- #
# TXT — structured, LLM-readable plain-text blocks (distinct from raw JSON)
# --------------------------------------------------------------------------- #
class TXTWriter(BaseWriter):
    """Human- and LLM-readable structured text. Each video becomes a clearly
    delimited block with labeled fields — designed so an LLM performing
    RAG/summarization over this file can parse it without JSON syntax
    getting in the way of the actual content, and so a human skimming it
    in a terminal can follow along easily.
    """

    format_name = "txt"

    _PREFERRED_ORDER = (
        "video_id", "status", "error_message", "channel_name", "channel_id",
        "views", "likes", "dislikes", "comment_count", "duration", "upload_date",
        "release_date", "language", "is_live", "is_age_restricted", "tags",
        "categories", "thumbnail_url", "share_link", "webpage_url",
        "playlist_title", "playlist_index", "description",
    )
    _SKIP_TRAILING = ("extracted_at", "source_url")

    def __init__(self, output_path: str | Path, fields: Optional[list[str]] = None):
        super().__init__(output_path, fields)
        self._fh: Optional[TextIO] = None
        self._index = 0

    def open(self) -> None:
        super().open()
        try:
            self._fh = self.output_path.open("w", encoding="utf-8")
        except OSError as exc:
            raise OutputWriterError(f"Could not open TXT output '{self.output_path}': {exc}") from exc

    @staticmethod
    def _fmt_value(value: Any) -> str:
        if value is None:
            return "N/A"
        if isinstance(value, list):
            return ", ".join(str(v) for v in value) if value else "N/A"
        return str(value)

    @staticmethod
    def _fmt_timestamp(seconds: Any) -> Optional[str]:
        """Render a raw seconds value (int/float, possibly None) as MM:SS
        or H:MM:SS for human/LLM readability. Returns None (not "N/A") for
        a genuinely absent value, since the caller uses None to decide
        whether to render an open-ended chapter ("12:30+") vs a closed one.
        """
        if seconds is None:
            return None
        try:
            total = int(round(float(seconds)))
        except (TypeError, ValueError):
            return None
        hours, remainder = divmod(total, 3600)
        minutes, secs = divmod(remainder, 60)
        if hours:
            return f"{hours}:{minutes:02d}:{secs:02d}"
        return f"{minutes}:{secs:02d}"

    def _render_block(self, record: VideoRecord) -> str:
        d = record.to_dict(self.fields)
        comments = d.pop("comments", [])
        chapters = d.pop("chapters", [])
        lines = [f"{'=' * 70}", f"VIDEO #{self._index}: {d.get('title') or '(untitled)'}", f"{'=' * 70}"]

        for key in self._PREFERRED_ORDER:
            if key in d:
                label = key.replace("_", " ").title()
                lines.append(f"{label}: {self._fmt_value(d[key])}")
        # Any remaining selected fields not covered above (forward-compat
        # if AVAILABLE_FIELDS grows without this list being updated).
        for key, value in d.items():
            if key in self._PREFERRED_ORDER or key in self._SKIP_TRAILING:
                continue
            label = key.replace("_", " ").title()
            lines.append(f"{label}: {self._fmt_value(value)}")

        if chapters:
            lines.append("")
            lines.append(f"--- Chapters ({len(chapters)}) ---")
            for ch in chapters:
                start = self._fmt_timestamp(ch.get("start_time"))
                end = self._fmt_timestamp(ch.get("end_time"))
                title = ch.get("title") or "(untitled chapter)"
                span = f"{start}-{end}" if end is not None else f"{start}+"
                lines.append(f"  [{span}] {title}")

        if comments:
            lines.append("")
            lines.append(f"--- Top Comments ({len(comments)} shown) ---")
            for c in comments:
                author = c.get("author") or "Unknown"
                text = (c.get("text") or "").replace("\n", " ")
                likes = c.get("like_count")
                lines.append(f"  [{author}] ({likes if likes is not None else '?'} likes): {text}")
                for reply in c.get("replies", []):
                    r_author = reply.get("author") or "Unknown"
                    r_text = (reply.get("text") or "").replace("\n", " ")
                    lines.append(f"    -> [{r_author}]: {r_text}")

        lines.append("")
        return "\n".join(lines)

    def write_batch(self, records: list[VideoRecord]) -> None:
        if self._fh is None:
            raise OutputWriterError("TXTWriter.write_batch called before open()")
        for record in records:
            self._index += 1
            self._fh.write(self._render_block(record))
            self._fh.write("\n")
        self._fh.flush()

    def close(self) -> None:
        if self._fh is not None:
            self._fh.close()
            self._fh = None
        super().close()


# --------------------------------------------------------------------------- #
# Registry + coordinator
# --------------------------------------------------------------------------- #
WRITER_REGISTRY: dict[str, type[BaseWriter]] = {
    "json": JSONWriter,
    "jsonl": JSONLWriter,
    "csv": CSVWriter,
    "xlsx": XLSXWriter,
    "txt": TXTWriter,
}


class ChunkedOutputCoordinator:
    """Owns every active writer (one per --format, plus DBWriter handled
    separately since it isn't a BaseWriter subclass) and fans out each
    completed chunk to all of them. This is the single place the CLI calls
    into — it never talks to individual writers directly.
    """

    def __init__(
        self,
        formats: list[str],
        output_dir: str | Path,
        base_filename: str,
        fields: Optional[list[str]] = None,
        db_writer: Any = None,  # Optional[db_manager.DBWriter]; kept as Any to avoid a circular import
    ):
        self.output_dir = Path(output_dir)
        self.fields = fields
        self.db_writer = db_writer
        self._writers: list[BaseWriter] = []
        for fmt in formats:
            if fmt == "db":
                continue  # handled via db_writer, not a BaseWriter
            writer_cls = WRITER_REGISTRY.get(fmt)
            if writer_cls is None:
                raise OutputWriterError(f"Unsupported output format: '{fmt}'")
            path = self.output_dir / f"{base_filename}.{fmt}"
            self._writers.append(writer_cls(path, fields=fields))

    def open(self) -> None:
        for w in self._writers:
            w.open()

    def write_chunk(self, records: list[VideoRecord]) -> None:
        for w in self._writers:
            try:
                w.write_batch(records)
            except OutputWriterError:
                raise
            except Exception as exc:  # noqa: BLE001 - one writer's bug must not corrupt/skip the others
                raise OutputWriterError(
                    f"Writer '{w.format_name}' failed on a batch of {len(records)} record(s): {exc}"
                ) from exc
        if self.db_writer is not None:
            self.db_writer.write_batch(records)

    def close(self) -> None:
        for w in self._writers:
            w.close()
        if self.db_writer is not None:
            self.db_writer.close()

    def __enter__(self) -> "ChunkedOutputCoordinator":
        self.open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.close()

    @property
    def output_paths(self) -> list[Path]:
        paths = [w.output_path for w in self._writers]
        if self.db_writer is not None:
            paths.append(self.db_writer.db_path)
        return paths
